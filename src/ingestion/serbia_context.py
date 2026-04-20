"""Canonical Serbia-context normalization for ingestion and RAG preparation."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.schemas.serbia_context import (
    SerbiaCanonicalIngestionBundle,
    SerbiaContextRecord,
    SerbiaDocumentRegistrationCandidate,
    SerbiaIngestionReadiness,
    SerbiaRecordProvenance,
    SerbiaSourceFamily,
    SerbiaStructuredContextRecord,
    SerbiaUrlKind,
)
from src.schemas.source_metadata import SourceType


def _slugify(value: str) -> str:
    """Return a filesystem- and id-safe slug."""

    lowered = value.strip().lower()
    lowered = re.sub(r"[^a-z0-9]+", "-", lowered)
    lowered = re.sub(r"-{2,}", "-", lowered)
    return lowered.strip("-")


def _canonical_id(
    source_family: SerbiaSourceFamily,
    *,
    title: str,
    source_file: str,
    source_row_number: int,
    extra: str = "",
) -> str:
    """Build a stable canonical id from family/title/provenance."""

    slug = _slugify(title)[:64] or "record"
    payload = f"{source_family}|{source_file}|{source_row_number}|{title}|{extra}"
    digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:10]
    return f"{source_family}-{slug}-{digest}"


def _normalize_header(value: Any) -> str:
    """Normalize dataset column headers for robust field lookup."""

    return re.sub(r"\s+", " ", str(value or "").strip())


def _as_text(value: Any) -> str:
    """Return a stripped string representation."""

    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _as_float(value: Any) -> float | None:
    """Parse numeric values that may use separators or symbols."""

    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _as_text(value)
    if not text:
        return None
    cleaned = text.replace(",", "").replace(" ", "").replace("EUR", "").replace("RSD", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _as_bool_yn(value: Any) -> bool | None:
    """Normalize Y/N-like fields to booleans."""

    text = _as_text(value).lower()
    if text in {"y", "yes", "true", "1"}:
        return True
    if text in {"n", "no", "false", "0"}:
        return False
    return None


def _as_iso_date(value: Any) -> str | None:
    """Normalize common date values to ISO strings."""

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _as_text(value)
    if not text:
        return None

    for pattern in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def classify_url_kind(url: str | None) -> SerbiaUrlKind:
    """Classify URLs without requiring direct-download normalization."""

    text = _as_text(url)
    if not text:
        return "unknown"
    lowered = text.lower()

    if any(domain in lowered for domain in ("drive.google.com", "docs.google.com", "1drv.ms", "sharepoint.com")):
        return "cloud_drive"

    if any(ext in lowered for ext in (".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx")):
        return "office_doc"

    if any(ext in lowered for ext in (".zip", ".rar", ".7z", ".tar.gz", ".tgz", ".tar")):
        return "archive"

    if any(ext in lowered for ext in (".pdf", ".csv", ".txt", ".rtf")):
        return "direct_document"

    if any(key in lowered for key in ("project-detail", "technicalassistancegrants", "wbif.eu", "/project/")):
        return "landing_page"

    if any(key in lowered for key in ("download", "wpdmdl=", "raw=1", "attachment")):
        return "direct_document"

    return "unknown"


def _ingestion_readiness(source_family: SerbiaSourceFamily, source_url: str | None, url_kind: SerbiaUrlKind) -> SerbiaIngestionReadiness:
    """Return ingestion readiness from source family and URL state."""

    structured_families = {"local_project_record", "wbif_project_record", "wbif_ta_record"}
    if not _as_text(source_url):
        if source_family in structured_families:
            return "metadata_only"
        return "missing_url"

    if url_kind == "direct_document":
        return "ready"
    if source_family in structured_families:
        return "metadata_only"
    return "needs_resolver"


def _url_roles(source_url: str | None, url_kind: SerbiaUrlKind) -> tuple[str | None, str | None]:
    """Split URLs into document and landing-page roles."""

    text = _as_text(source_url) or None
    if text is None:
        return None, None
    if url_kind == "direct_document":
        return text, None
    if url_kind == "landing_page":
        return None, text
    return None, None


def _infer_sector_tags(*values: str) -> list[str]:
    """Infer sector tags from title/description text."""

    text = " ".join(value.lower() for value in values if value).strip()
    if not text:
        return []
    tags: list[str] = []
    rules = {
        "environment": ("air", "waste", "climate", "biodiversity", "water", "emission"),
        "transport": ("transport", "rail", "road", "highway", "bicycle"),
        "energy": ("energy", "gas", "electricity", "renewable", "efficiency"),
        "digitalization": ("digital", "broadband", "data centre", "ict"),
        "social": ("housing", "education", "health", "social"),
    }
    for tag, keywords in rules.items():
        if any(keyword in text for keyword in keywords):
            tags.append(tag)
    return tags


def _municipality_id_from_name(name: str | None) -> str | None:
    """Create a stable municipality id candidate for Serbia rows."""

    text = _as_text(name)
    if not text:
        return None
    return f"srb-{_slugify(text)}"


def _structured_searchable_text(record: SerbiaContextRecord) -> str:
    """Build a retrieval-oriented text representation for structured records."""

    parts = [
        record.display_title,
        record.title,
        record.municipality_name or "",
        record.district_name or "",
        record.region_name or "",
        ", ".join(record.category_tags),
        ", ".join(record.sector_tags),
    ]
    for key, value in sorted(record.attributes.items()):
        parts.append(f"{key}: {value}")
    return "\n".join(part for part in parts if part).strip()


def _build_record(
    *,
    source_family: SerbiaSourceFamily,
    title: str,
    display_title: str,
    source_file: str,
    source_sheet: str | None,
    source_row_number: int,
    source_url: str | None,
    municipality_name: str | None = None,
    municipality_code: str | None = None,
    district_name: str | None = None,
    region_name: str | None = None,
    category_tags: list[str] | None = None,
    sector_tags: list[str] | None = None,
    attributes: dict[str, object] | None = None,
    extra_id_seed: str = "",
) -> SerbiaContextRecord:
    """Create one canonical record with URL role/classification fields."""

    normalized_url = _as_text(source_url) or None
    url_kind = classify_url_kind(normalized_url)
    readiness = _ingestion_readiness(source_family, normalized_url, url_kind)
    document_url, landing_page_url = _url_roles(normalized_url, url_kind)

    record = SerbiaContextRecord(
        canonical_id=_canonical_id(
            source_family,
            title=title,
            source_file=source_file,
            source_row_number=source_row_number,
            extra=extra_id_seed,
        ),
        source_family=source_family,
        title=title,
        display_title=display_title,
        municipality_name=municipality_name,
        municipality_code=municipality_code,
        district_name=district_name,
        region_name=region_name,
        category_tags=category_tags or [],
        sector_tags=sector_tags or [],
        source_url=normalized_url,
        document_url=document_url,
        landing_page_url=landing_page_url,
        url_kind=url_kind,
        ingestion_readiness=readiness,
        provenance=SerbiaRecordProvenance(
            source_file=source_file,
            source_sheet=source_sheet,
            source_row_number=source_row_number,
        ),
        summary_text="",
        attributes=attributes or {},
    )
    return record.model_copy(update={"summary_text": _structured_searchable_text(record)})


def normalize_national_policy_records(path: Path) -> list[SerbiaContextRecord]:
    """Normalize the national strategy/policy workbook."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=1, values_only=True)
    header = [_normalize_header(value) for value in next(rows)]
    index = {name: idx for idx, name in enumerate(header)}

    records: list[SerbiaContextRecord] = []
    for row_number, row in enumerate(rows, start=2):
        title = _as_text(row[index.get("Document", 0)] if "Document" in index else row[0])
        if not title:
            continue
        source_url = _as_text(row[index.get("URL", 1)] if "URL" in index and len(row) > index["URL"] else "")
        sector_tags = _infer_sector_tags(title)
        records.append(
            _build_record(
                source_family="national_policy_document",
                title=title,
                display_title=f"Serbia National Policy: {title}",
                source_file=path.name,
                source_sheet=sheet.title,
                source_row_number=row_number,
                source_url=source_url,
                category_tags=["serbia", "national-policy"],
                sector_tags=sector_tags,
                attributes={"document_name": title},
            )
        )
    return records


def normalize_municipal_development_plan_records(path: Path) -> list[SerbiaContextRecord]:
    """Normalize municipal development plan rows."""

    records: list[SerbiaContextRecord] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            municipality_name = _as_text(row.get("ENGLISH_NAME"))
            district_name = _as_text(row.get("NAME_1")) or None
            municipality_code = _as_text(row.get("GID_2")) or None
            source_url = _as_text(row.get("download_link"))
            title = f"{municipality_name} Local Development Plan" if municipality_name else "Municipal Development Plan"
            records.append(
                _build_record(
                    source_family="municipal_development_plan",
                    title=title,
                    display_title=title,
                    source_file=path.name,
                    source_sheet=None,
                    source_row_number=row_number,
                    source_url=source_url,
                    municipality_name=municipality_name or None,
                    municipality_code=municipality_code,
                    district_name=district_name,
                    category_tags=["serbia", "municipal-policy", "development-plan"],
                    sector_tags=_infer_sector_tags(title),
                    attributes={
                        "name_1": _as_text(row.get("NAME_1")),
                        "gid_2": _as_text(row.get("GID_2")),
                        "english_name": municipality_name,
                        "download_link": source_url,
                    },
                    extra_id_seed=municipality_code or municipality_name,
                )
            )
    return records


def normalize_local_project_records(path: Path) -> list[SerbiaContextRecord]:
    """Normalize local infrastructure/institutional project workbook rows."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=1, values_only=True)
    header = [_normalize_header(value) for value in next(rows)]
    columns = {name: idx for idx, name in enumerate(header)}

    sector_code_map = {
        "T": "transport",
        "M": "mobility",
        "E": "energy",
        "W": "water",
        "S": "social",
    }

    def value_for(row: tuple[Any, ...], column: str) -> Any:
        """Return a column value from a normalized header map."""

        idx = columns.get(column)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records: list[SerbiaContextRecord] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(_as_text(value) for value in row):
            continue
        municipality = _as_text(value_for(row, "LSG")) or None
        project_title = _as_text(value_for(row, "INVESTMENT PROJECT TITLE"))
        if not project_title:
            continue
        sector_code = _as_text(value_for(row, "PROJECT SECTOR")).upper()
        sector_tags = []
        if sector_code:
            sector_tags.append(sector_code_map.get(sector_code, f"sector-{sector_code.lower()}"))
            sector_tags.append(f"sector-code-{sector_code.lower()}")
        sector_tags.extend(_infer_sector_tags(project_title))

        records.append(
            _build_record(
                source_family="local_project_record",
                title=project_title,
                display_title=f"{municipality}: {project_title}" if municipality else project_title,
                source_file=path.name,
                source_sheet=sheet.title,
                source_row_number=row_number,
                source_url=None,
                municipality_name=municipality,
                region_name=_as_text(value_for(row, "REGION")) or None,
                category_tags=["serbia", "local-project"],
                sector_tags=sorted(set(tag for tag in sector_tags if tag)),
                attributes={
                    "lsg": municipality or "",
                    "total_allocated_budget_eur": _as_float(value_for(row, "TOTAL ALLOCATED BUDGET PER LSG EUR")),
                    "investment_amount_eur": _as_float(value_for(row, "INVESTMENT AMOUNT OF PROPOSED PROJECT [EUR]")),
                    "investment_amount_rsd": _as_float(value_for(row, "INVESTMENT AMOUNT OF PROPOSED PROJECT [RSD]")),
                    "region": _as_text(value_for(row, "REGION")),
                    "group": _as_text(value_for(row, "GROUP")),
                    "project_sector_code": sector_code,
                    "decision_signed": _as_bool_yn(value_for(row, "DECISION OF MCTI SIGNED Y/N")),
                    "screening_report_or_sep": _as_bool_yn(value_for(row, "SCREENING REPORT / SEP Y/N")),
                    "ga_signed": _as_bool_yn(value_for(row, "GA SIGNED Y/N")),
                    "ga_signature_date": _as_iso_date(value_for(row, "DATE OF GA SIGNATURE")),
                    "ga_group": _as_text(value_for(row, "GA GROUP")),
                    "internal_grant_agreement_number": _as_text(value_for(row, "INTERNAL GRANT AGREEMENT NUMBER")),
                    "grant_agreement_value_eur": _as_float(value_for(row, "GRANT AGREEMENT VALUE [EUR]")),
                    "exchange_rate": _as_float(value_for(row, "EXCHANGE RATE")),
                    "converted_grant_value_rsd": _as_float(value_for(row, "CONVERTED GRANT AGREEMENT VALUE [RSD]")),
                    "allocated_budget_ratio": _as_float(value_for(row, "% OF TOTAL ALLOCATED LSG BUDGET")),
                    "lsg_cofinancing": _as_bool_yn(value_for(row, "LSG COFINANCING Y/N")),
                },
                extra_id_seed=f"{municipality}|{project_title}",
            )
        )
    return records


def normalize_wbif_project_records(path: Path) -> list[SerbiaContextRecord]:
    """Normalize and filter WBIF investment project rows for Serbia context."""

    records: list[SerbiaContextRecord] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            beneficiaries = _as_text(row.get("Beneficiary Country"))
            if "serb" not in beneficiaries.lower():
                continue

            title = _as_text(row.get("Project Name")) or "WBIF Project"
            source_url = _as_text(row.get("Project WBIF URL"))
            sector = _as_text(row.get("Investment Sector"))
            records.append(
                _build_record(
                    source_family="wbif_project_record",
                    title=title,
                    display_title=title,
                    source_file=path.name,
                    source_sheet=None,
                    source_row_number=row_number,
                    source_url=source_url,
                    category_tags=["serbia", "wbif-project"],
                    sector_tags=sorted(
                        set(
                            tag
                            for tag in [*([_slugify(sector)] if sector else []), *_infer_sector_tags(title, sector)]
                            if tag
                        )
                    ),
                    attributes={
                        "project_wbif_url": source_url,
                        "project_description": _as_text(row.get("Project Description")),
                        "beneficiary_country": beneficiaries,
                        "investment_sector": sector,
                        "estimated_completion": _as_iso_date(row.get("Estimated Completion")),
                        "beneficiary_body": _as_text(row.get("Beneficiary Body")),
                        "total_financing": _as_float(row.get("Total Financing")),
                        "total_grant": _as_float(row.get("Total Grant")),
                        "total_loan": _as_float(row.get("Total Loan")),
                        "project_benefits": _as_text(row.get("Project Benefits")),
                    },
                    extra_id_seed=f"{beneficiaries}|{_as_text(row.get('Project WBIF URL'))}",
                )
            )
    return records


def normalize_wbif_ta_records(path: Path) -> list[SerbiaContextRecord]:
    """Normalize and filter WBIF TA rows for Serbia context."""

    records: list[SerbiaContextRecord] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            beneficiaries = _as_text(row.get("Beneficiaries"))
            project_name = _as_text(row.get("project_name"))
            if "serb" not in beneficiaries.lower() and "serb" not in project_name.lower():
                continue

            source_url = _as_text(row.get("TA URL"))
            sector = _as_text(row.get("Sector"))
            status = _as_text(row.get("Status"))
            records.append(
                _build_record(
                    source_family="wbif_ta_record",
                    title=project_name or "WBIF Technical Assistance",
                    display_title=project_name or "WBIF Technical Assistance",
                    source_file=path.name,
                    source_sheet=None,
                    source_row_number=row_number,
                    source_url=source_url,
                    category_tags=["serbia", "wbif-ta"],
                    sector_tags=sorted(
                        set(
                            tag
                            for tag in [*([_slugify(sector)] if sector else []), *_infer_sector_tags(project_name, sector)]
                            if tag
                        )
                    ),
                    attributes={
                        "ta_url": source_url,
                        "ta_description": _as_text(row.get("TA Description")),
                        "beneficiaries": beneficiaries,
                        "lead_ifi": _as_text(row.get("Lead IFI")),
                        "sector": sector,
                        "type": _as_text(row.get("Type")),
                        "status": _slugify(status) if status else "",
                        "date_of_award": _as_iso_date(row.get("Date of award")),
                        "date_of_completion": _as_iso_date(row.get("Date of completion")),
                        "total_wbif_grant": _as_float(row.get("Total WBIF grant")),
                        "beneficiary": _as_text(row.get("Beneficiary")),
                        "project_code": _as_text(row.get("Project Code")),
                    },
                    extra_id_seed=f"{project_name}|{_as_text(row.get('Project Code'))}",
                )
            )
    return records


def build_document_registration_candidates(records: list[SerbiaContextRecord]) -> list[SerbiaDocumentRegistrationCandidate]:
    """Build document registration candidates from canonical records."""

    source_type_map: dict[SerbiaSourceFamily, SourceType] = {
        "national_policy_document": "policy_document",
        "municipal_development_plan": "municipal_development_plan",
        "local_project_record": "project_document",
        "wbif_project_record": "project_page",
        "wbif_ta_record": "project_page",
    }
    document_like_kinds = {"direct_document", "cloud_drive", "office_doc"}

    candidates: list[SerbiaDocumentRegistrationCandidate] = []
    for record in records:
        if record.url_kind not in document_like_kinds:
            continue
        uri = record.document_url or record.source_url
        if not uri:
            continue
        candidates.append(
            SerbiaDocumentRegistrationCandidate(
                canonical_id=record.canonical_id,
                source_family=record.source_family,
                source_type=source_type_map[record.source_family],
                title=record.display_title,
                uri=uri,
                municipality_name=record.municipality_name,
                source_url=record.source_url,
                document_url=record.document_url,
                landing_page_url=record.landing_page_url,
                url_kind=record.url_kind,
                ingestion_readiness=record.ingestion_readiness,
                municipality_id=_municipality_id_from_name(record.municipality_name),
                category=record.sector_tags[0] if record.sector_tags else None,
                normalized_metadata={
                    "canonical_id": record.canonical_id,
                    "source_family": record.source_family,
                    "country_code": record.country_code,
                    "provenance_file": record.provenance.source_file,
                    "provenance_row": str(record.provenance.source_row_number),
                },
            )
        )
    return candidates


def build_structured_context_records(records: list[SerbiaContextRecord]) -> list[SerbiaStructuredContextRecord]:
    """Build structured row context for embedding/RAG without requiring fetch/crawl."""

    return [
        SerbiaStructuredContextRecord(
            canonical_id=record.canonical_id,
            source_family=record.source_family,
            title=record.display_title,
            municipality_name=record.municipality_name,
            district_name=record.district_name,
            region_name=record.region_name,
            category_tags=record.category_tags,
            sector_tags=record.sector_tags,
            searchable_text=record.summary_text,
            attributes=record.attributes,
            provenance=record.provenance,
        )
        for record in records
    ]


def build_serbia_canonical_context_bundle(data_dir: Path) -> SerbiaCanonicalIngestionBundle:
    """Build canonical Serbia context records and derived ingestion views."""

    national_path = data_dir / "national_strategy_policies_law.xlsx"
    municipal_path = data_dir / "serbia_local_dev_plans_final.csv"
    local_projects_path = data_dir / "serbia_lsg_projects.xlsx"
    wbif_projects_path = data_dir / "wbif_projects.csv"
    wbif_tas_path = data_dir / "wbif_TAs.csv"

    records: list[SerbiaContextRecord] = []
    records.extend(normalize_national_policy_records(national_path))
    records.extend(normalize_municipal_development_plan_records(municipal_path))
    records.extend(normalize_local_project_records(local_projects_path))
    records.extend(normalize_wbif_project_records(wbif_projects_path))
    records.extend(normalize_wbif_ta_records(wbif_tas_path))

    document_candidates = build_document_registration_candidates(records)
    structured_records = build_structured_context_records(records)

    stats = {
        "records_total": len(records),
        "national_policy_document": len([item for item in records if item.source_family == "national_policy_document"]),
        "municipal_development_plan": len([item for item in records if item.source_family == "municipal_development_plan"]),
        "local_project_record": len([item for item in records if item.source_family == "local_project_record"]),
        "wbif_project_record": len([item for item in records if item.source_family == "wbif_project_record"]),
        "wbif_ta_record": len([item for item in records if item.source_family == "wbif_ta_record"]),
        "missing_url_records": len([item for item in records if item.ingestion_readiness == "missing_url"]),
        "document_registration_candidates": len(document_candidates),
        "structured_context_records": len(structured_records),
    }
    return SerbiaCanonicalIngestionBundle(
        records=records,
        document_registration_candidates=document_candidates,
        structured_context_records=structured_records,
        stats=stats,
    )


def export_serbia_canonical_context_bundle(bundle: SerbiaCanonicalIngestionBundle, output_dir: Path) -> None:
    """Export canonical Serbia context records and derived views to JSON files."""

    output_dir.mkdir(parents=True, exist_ok=True)
    records_path = output_dir / "serbia_context_records.jsonl"
    document_candidates_path = output_dir / "serbia_document_registration_candidates.jsonl"
    structured_records_path = output_dir / "serbia_structured_context_records.jsonl"
    stats_path = output_dir / "serbia_context_stats.json"

    with records_path.open("w", encoding="utf-8") as handle:
        for record in bundle.records:
            handle.write(json.dumps(record.model_dump(mode="json"), ensure_ascii=True) + "\n")

    with document_candidates_path.open("w", encoding="utf-8") as handle:
        for record in bundle.document_registration_candidates:
            handle.write(json.dumps(record.model_dump(mode="json"), ensure_ascii=True) + "\n")

    with structured_records_path.open("w", encoding="utf-8") as handle:
        for record in bundle.structured_context_records:
            handle.write(json.dumps(record.model_dump(mode="json"), ensure_ascii=True) + "\n")

    with stats_path.open("w", encoding="utf-8") as handle:
        json.dump(bundle.stats, handle, ensure_ascii=True, indent=2, sort_keys=True)
